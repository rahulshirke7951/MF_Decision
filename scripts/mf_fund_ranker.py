"""
MF Fund Ranker — Decision Engine
=================================
Reads your dashboard_data.xlsx, scores every scheme within its category,
and outputs a ranked Excel file: mf_ranked_screener.xlsx

Scoring Formula:
  - 1Y Return   : 40%
  - 3Y CAGR     : 40%
  - 1M Return   : 7%   (short-term outlook)
  - 3M Return   : 7%   (short-term outlook)
  - 6M Return   : 6%   (short-term outlook)

Usage:
  pip install pandas openpyxl
  python mf_fund_ranker.py

Place dashboard_data.xlsx in the same folder as this script.
Output: mf_ranked_screener.xlsx
"""

import pandas as pd
import os
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── CONFIG ──────────────────────────────────────────────────────────────────
INPUT_FILE  = "dashboard_data.xlsx"
OUTPUT_FILE = "mf_ranked_screener.xlsx"

WEIGHTS = {
    "return_1y":  0.40,
    "return_3y":  0.40,
    "return_1m":  0.07,   # 1-month short-term outlook
    "return_3m":  0.07,   # 3-month short-term outlook
    "return_6m":  0.06,   # 6-month short-term outlook
}

# ── FILTERS ─────────────────────────────────────────────────────────────────
# Set a value to None to skip that filter entirely.
# Column names must match exactly what is in your Excel file (case-insensitive).
FILTERS = {
    "cat_level_1": "Open Ended Schemes",
    "cat_level_2": "Equity Scheme",
    "plan_type":   "Regular",
    "option_type": "Growth",
}

# ── COLUMN AUTO-DETECTION ───────────────────────────────────────────────────
COLUMN_ALIASES = {
    "scheme_name": [
        "scheme_name", "scheme name", "fund name", "name", "schemename",
        "fund", "scheme"
    ],
    "category": [
        "cat_level_3", "cat_level_2", "cat_level_1",
        "category", "scheme_category", "scheme category", "cat",
        "fund_category", "fund category", "sub_category", "sub category",
        "category_name", "sub_type"
    ],
    "amc": [
        "amc", "amc_name", "amc name", "fund_house", "fund house",
        "amcname", "asset_management_company", "house"
    ],
    "return_1m": [
        "return_30d", "return_1m", "1m_return", "1month_return", "ret_1m",
        "trailing_1m", "1m return", "1 month return"
    ],
    "return_3m": [
        "return_90d", "return_3m", "3m_return", "3month_return", "ret_3m",
        "trailing_3m", "3m return", "3 month return"
    ],
    "return_6m": [
        "return_180d", "return_6m", "6m_return", "6month_return", "ret_6m",
        "trailing_6m", "6m return", "6 month return"
    ],
    "return_1y": [
        "return_365d", "return_1y", "1y_return", "1yr_return", "returns_1y",
        "1y return", "1 year return", "1yr", "ret_1y", "cagr_1y", "1y_cagr",
        "trailing_1y", "1year_return", "ann_return_1y"
    ],
    "return_3y": [
        "return_1095d", "return_3y", "3y_return", "3yr_return", "returns_3y",
        "3y return", "3 year return", "3yr", "ret_3y", "cagr_3y", "3y_cagr",
        "trailing_3y", "3year_return", "ann_return_3y"
    ],
    "nav": [
        "nav", "net_asset_value", "current_nav", "latest_nav"
    ],
    "plan": [
        "plan", "plan_type", "scheme_plan", "direct_regular"
    ],
}

COLORS = {
    "header_bg":   "1A3A5C",   # dark navy
    "header_fg":   "FFFFFF",
    "rank1_bg":    "FFD700",   # gold
    "rank2_bg":    "E8E8E8",   # silver
    "rank3_bg":    "D4956A",   # bronze
    "positive":    "1E7A4B",   # dark green
    "negative":    "C0392B",   # red
    "cat_header":  "2D6A8A",   # teal
    "cat_fg":      "FFFFFF",
    "alt_row":     "F5F0E8",   # cream
    "score_bg":    "EBF5FB",
    "border":      "CCCCCC",
}


def detect_column(df_cols, key):
    """Find the actual column name from aliases."""
    cols_lower = {c.lower().strip(): c for c in df_cols}
    for alias in COLUMN_ALIASES.get(key, []):
        if alias.lower() in cols_lower:
            return cols_lower[alias.lower()]
    return None


def apply_filters(df):
    """
    Apply FILTERS to the DataFrame.
    Each key in FILTERS must match a column name in the data (case-insensitive).
    Values are matched case-insensitively; set a filter value to None to skip it.
    """
    cols_lower = {c.lower().strip(): c for c in df.columns}
    active_filters = {k: v for k, v in FILTERS.items() if v is not None}

    if not active_filters:
        print("   ℹ️  No filters configured — using all rows.")
        return df

    print(f"\n🔎 Applying filters ({len(active_filters)} active):")
    for col_key, value in active_filters.items():
        actual_col = cols_lower.get(col_key.lower().strip())
        if actual_col is None:
            print(f"   ⚠️  Filter column '{col_key}' not found in data — skipping this filter.")
            continue
        before = len(df)
        df = df[df[actual_col].astype(str).str.strip().str.lower() == str(value).strip().lower()]
        print(f"   ✅ {col_key} = '{value}'  →  {before} → {len(df)} rows")

    print(f"   📋 Rows after all filters: {len(df)}")
    return df


def load_data(filepath):
    """Load all sheets, combine, apply filters, and return a unified DataFrame."""
    print(f"\n📂 Reading: {filepath}")
    sheets = pd.read_excel(filepath, sheet_name=None)
    print(f"   Sheets found: {list(sheets.keys())}")

    frames = []
    for name, df in sheets.items():
        df["_source_sheet"] = name
        frames.append(df)

    combined = pd.concat(frames, ignore_index=True)
    print(f"   Total rows across all sheets: {len(combined)}")
    print(f"   Columns: {list(combined.columns)}")

    combined = apply_filters(combined)
    return combined


def map_columns(df):
    """Detect and map standard field names."""
    mapping = {}
    for key in COLUMN_ALIASES:
        col = detect_column(df.columns, key)
        mapping[key] = col
        status = f"✅ '{col}'" if col else "❌ not found"
        print(f"   {key:<15} → {status}")
    return mapping


def to_numeric(series):
    """Clean and convert a series to numeric, coercing errors."""
    if series is None:
        return pd.Series(dtype=float)
    s = series.astype(str).str.replace('%', '').str.replace(',', '').str.strip()
    return pd.to_numeric(s, errors='coerce')


def percentile_score(series):
    """Rank-based percentile score 0–100 within the group."""
    ranks = series.rank(method='min', na_option='bottom')
    return (ranks - 1) / max(len(series) - 1, 1) * 100


def score_funds(df, mapping):
    """Compute composite score for each fund row."""
    df = df.copy()

    # Pull numeric columns
    r1m = to_numeric(df[mapping["return_1m"]] if mapping["return_1m"] else None)
    r3m = to_numeric(df[mapping["return_3m"]] if mapping["return_3m"] else None)
    r6m = to_numeric(df[mapping["return_6m"]] if mapping["return_6m"] else None)
    r1  = to_numeric(df[mapping["return_1y"]] if mapping["return_1y"] else None)
    r3  = to_numeric(df[mapping["return_3y"]] if mapping["return_3y"] else None)

    df["_r1m"] = r1m
    df["_r3m"] = r3m
    df["_r6m"] = r6m
    df["_r1"]  = r1
    df["_r3"]  = r3

    cat_col = mapping["category"]
    if not cat_col:
        df["_category_clean"] = "All Funds"
    else:
        df["_category_clean"] = df[cat_col].astype(str).str.strip().str.title()

    # Determine which columns actually have data
    has_r1m = mapping["return_1m"] and df["_r1m"].notna().any()
    has_r3m = mapping["return_3m"] and df["_r3m"].notna().any()
    has_r6m = mapping["return_6m"] and df["_r6m"].notna().any()
    has_r1  = mapping["return_1y"] and df["_r1"].notna().any()
    has_r3  = mapping["return_3y"] and df["_r3"].notna().any()

    # Redistribute weights if some columns are missing
    active = {k: v for k, v in [
        ("return_1y", has_r1), ("return_3y", has_r3),
        ("return_1m", has_r1m), ("return_3m", has_r3m), ("return_6m", has_r6m),
    ] if v}
    raw_total = sum(WEIGHTS[k] for k in active)
    adj_weights = {k: WEIGHTS[k] / raw_total for k in active}
    print(f"\n   Active scoring columns: {list(active.keys())}")
    print(f"   Adjusted weights: { {k: f'{v:.0%}' for k, v in adj_weights.items()} }")

    # Score per category
    score_parts = []
    for cat, grp in df.groupby("_category_clean"):
        g = grp.copy()
        score = pd.Series(0.0, index=g.index)
        if has_r1m:
            score += percentile_score(g["_r1m"].fillna(g["_r1m"].median())) * adj_weights["return_1m"]
        if has_r3m:
            score += percentile_score(g["_r3m"].fillna(g["_r3m"].median())) * adj_weights["return_3m"]
        if has_r6m:
            score += percentile_score(g["_r6m"].fillna(g["_r6m"].median())) * adj_weights["return_6m"]
        if has_r1:
            score += percentile_score(g["_r1"].fillna(g["_r1"].median())) * adj_weights["return_1y"]
        if has_r3:
            score += percentile_score(g["_r3"].fillna(g["_r3"].median())) * adj_weights["return_3y"]
        g["_composite_score"] = score.round(1)
        g["_rank"] = g["_composite_score"].rank(method='min', ascending=False).astype(int)
        score_parts.append(g)

    return pd.concat(score_parts, ignore_index=True)


def get_col_val(row, col):
    return row[col] if col and col in row.index else None


def build_excel(df, mapping, output_path):
    """Write the ranked screener to a well-formatted Excel workbook."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    thin = Side(style='thin', color=COLORS["border"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_font(bold=True, color="FFFFFF", size=10):
        return Font(name="Arial", bold=bold, color=color, size=size)

    def cell_font(bold=False, color="000000", size=9):
        return Font(name="Arial", bold=bold, color=color, size=size)

    def fill(hex_color):
        return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

    def pct(val):
        if pd.isna(val): return "—"
        return f"{val:.1f}%"

    def score_color(score):
        if score >= 80: return "1E7A4B"
        if score >= 60: return "2980B9"
        if score >= 40: return "F39C12"
        return "C0392B"

    # Column layout:
    # 1=Rank, 2=Scheme Name, 3=AMC,
    # 4=1M Ret, 5=3M Ret, 6=6M Ret,   ← Short-Term group
    # 7=1Y Ret, 8=3Y CAGR,             ← Long-Term group
    # 9=Composite Score, 10=Signal, 11=Category
    RETURN_COLS = {4, 5, 6, 7, 8}
    SCORE_COL   = 9

    categories = sorted(df["_category_clean"].unique())
    print(f"\n📊 Writing {len(categories)} category sheets...")

    # ── Per-category sheets ──────────────────────────────────────────────
    for cat in categories:
        cat_df = df[df["_category_clean"] == cat].sort_values("_rank")
        safe_name = (cat[:31].replace("/", "-").replace("\\", "-")
                     .replace("*", "").replace("?", "")
                     .replace("[", "").replace("]", ""))
        ws = wb.create_sheet(title=safe_name)

        # Row 1 — Title
        ws.merge_cells("A1:K1")
        ws["A1"] = f"📈  {cat}  —  Fund Performance Ranker"
        ws["A1"].font = Font(name="Arial", bold=True, size=13, color=COLORS["cat_fg"])
        ws["A1"].fill = fill(COLORS["cat_header"])
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 28

        # Row 2 — Sub-header
        ws.merge_cells("A2:K2")
        ws["A2"] = (
            f"Scoring: 1Y (40%) + 3Y CAGR (40%) + Short-Term 1M/3M/6M (20%)  "
            f"|  Total funds: {len(cat_df)}"
        )
        ws["A2"].font = Font(name="Arial", italic=True, size=8, color="555555")
        ws["A2"].fill = fill("F0F4F8")
        ws["A2"].alignment = Alignment(horizontal="left", indent=1)
        ws.row_dimensions[2].height = 16

        # Row 3 — Column group labels
        for blank_col in [1, 2, 3, 9, 10, 11]:
            ws.cell(row=3, column=blank_col).fill = fill(COLORS["header_bg"])

        ws.merge_cells("D3:F3")
        ws["D3"] = "◀  Short-Term Outlook  ▶"
        ws["D3"].font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        ws["D3"].fill = fill("3D6B8A")
        ws["D3"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("G3:H3")
        ws["G3"] = "◀  Long-Term  ▶"
        ws["G3"].font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        ws["G3"].fill = fill(COLORS["header_bg"])
        ws["G3"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 14

        # Row 4 — Column headers
        headers = [
            "Rank", "Scheme Name", "AMC",
            "1M Return", "3M Return", "6M Return",
            "1Y Return", "3Y CAGR",
            "Composite Score", "Signal", "Category"
        ]
        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=hdr)
            cell.font = hdr_font()
            cell.fill = fill(COLORS["header_bg"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws.row_dimensions[4].height = 18

        # Data rows — start at row 5
        for i, (_, row) in enumerate(cat_df.iterrows(), 5):
            rank  = row.get("_rank", i - 4)
            name  = get_col_val(row, mapping["scheme_name"]) or "—"
            amc   = get_col_val(row, mapping["amc"])         or "—"
            r1m   = row.get("_r1m")
            r3m   = row.get("_r3m")
            r6m   = row.get("_r6m")
            r1    = row.get("_r1")
            r3    = row.get("_r3")
            score = row.get("_composite_score", 0)

            if score >= 75:   signal = "⭐ Strong Buy"
            elif score >= 55: signal = "✅ Buy"
            elif score >= 40: signal = "⚠️ Watch"
            else:             signal = "❌ Avoid"

            values = [
                rank, name, amc,
                pct(r1m), pct(r3m), pct(r6m),
                pct(r1), pct(r3),
                round(score, 1), signal, cat
            ]

            if rank == 1:    row_bg = COLORS["rank1_bg"]
            elif rank == 2:  row_bg = COLORS["rank2_bg"]
            elif rank == 3:  row_bg = COLORS["rank3_bg"]
            elif i % 2 == 0: row_bg = COLORS["alt_row"]
            else:             row_bg = "FFFFFF"

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=i, column=col_idx, value=val)
                cell.font = cell_font(bold=(rank <= 3), size=9)
                cell.fill = fill(row_bg)
                cell.border = border
                cell.alignment = Alignment(
                    horizontal="center", vertical="center",
                    wrap_text=(col_idx == 2)
                )
                if col_idx in RETURN_COLS and isinstance(val, str) and val != "—":
                    num = float(val.replace('%', ''))
                    cell.font = Font(
                        name="Arial", size=9, bold=(rank <= 3),
                        color=COLORS["positive"] if num > 0 else COLORS["negative"]
                    )
                if col_idx == SCORE_COL:
                    cell.font = Font(name="Arial", bold=True, size=9, color=score_color(score))

            ws.row_dimensions[i].height = 16

        widths = [6, 52, 20, 11, 11, 11, 11, 11, 16, 14, 22]
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.freeze_panes = "A5"
        print(f"   ✅ {cat} — {len(cat_df)} funds ranked")

    # ── Summary Sheet ────────────────────────────────────────────────────
    ws_sum = wb.create_sheet(title="🏆 SUMMARY", index=0)

    ws_sum.merge_cells("A1:J1")
    ws_sum["A1"] = "MF INTELLIGENCE — TOP FUND PER CATEGORY"
    ws_sum["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws_sum["A1"].fill = fill("0D1117")
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 30

    ws_sum.merge_cells("A2:J2")
    ws_sum["A2"] = (
        "Scoring: 1Y Returns (40%) + 3Y CAGR (40%) + Short-Term 1M/3M/6M (20%)  "
        "|  Category grouping: cat_level_3"
    )
    ws_sum["A2"].font = Font(name="Arial", italic=True, size=8, color="555555")
    ws_sum["A2"].fill = fill("F5F5F5")
    ws_sum["A2"].alignment = Alignment(horizontal="center")
    ws_sum.row_dimensions[2].height = 14

    sum_headers = [
        "Category", "#1 Fund", "AMC",
        "1M Return", "3M Return", "6M Return",
        "1Y Return", "3Y CAGR",
        "Score", "Signal"
    ]
    for ci, hdr in enumerate(sum_headers, 1):
        c = ws_sum.cell(row=3, column=ci, value=hdr)
        c.font = hdr_font()
        c.fill = fill(COLORS["header_bg"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws_sum.row_dimensions[3].height = 18

    SUM_RETURN_COLS = {4, 5, 6, 7, 8}
    SUM_SCORE_COL   = 9

    for i, cat in enumerate(categories, 4):
        cat_df = df[df["_category_clean"] == cat].sort_values("_rank")
        if cat_df.empty: continue
        top   = cat_df.iloc[0]
        name  = get_col_val(top, mapping["scheme_name"]) or "—"
        amc   = get_col_val(top, mapping["amc"])         or "—"
        r1m   = top.get("_r1m")
        r3m   = top.get("_r3m")
        r6m   = top.get("_r6m")
        r1    = top.get("_r1")
        r3    = top.get("_r3")
        score = top.get("_composite_score", 0)
        signal = "⭐ Strong Buy" if score >= 75 else "✅ Buy" if score >= 55 else "⚠️ Watch"

        row_bg = COLORS["alt_row"] if i % 2 == 0 else "FFFFFF"
        vals = [cat, name, amc, pct(r1m), pct(r3m), pct(r6m), pct(r1), pct(r3), round(score, 1), signal]
        for ci, v in enumerate(vals, 1):
            c = ws_sum.cell(row=i, column=ci, value=v)
            c.font = cell_font(size=9)
            c.fill = fill(row_bg)
            c.border = border
            c.alignment = Alignment(
                horizontal="center" if ci != 2 else "left",
                vertical="center", wrap_text=(ci == 2)
            )
            if ci in SUM_RETURN_COLS and isinstance(v, str) and v != "—":
                num = float(v.replace('%', ''))
                c.font = Font(
                    name="Arial", size=9,
                    color=COLORS["positive"] if num > 0 else COLORS["negative"]
                )
            if ci == SUM_SCORE_COL:
                c.font = Font(name="Arial", bold=True, size=9, color=score_color(score))
        ws_sum.row_dimensions[i].height = 18

    for ci, w in enumerate([28, 52, 20, 11, 11, 11, 11, 11, 10, 14], 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w
    ws_sum.freeze_panes = "A4"

    wb.save(output_path)
    print(f"\n✅ Output saved → {output_path}")


def main():
    if not os.path.exists(INPUT_FILE):
        print(f"\n❌ File not found: {INPUT_FILE}")
        print("   Place dashboard_data.xlsx in the same folder as this script and re-run.")
        return

    df = load_data(INPUT_FILE)

    if df.empty:
        print("\n❌ No rows remain after applying filters. Check your FILTERS config.")
        return

    print("\n🔍 Auto-detecting columns...")
    mapping = map_columns(df)

    missing_critical = [k for k in ("scheme_name", "return_1y", "return_3y") if not mapping[k]]
    if missing_critical:
        print(f"\n⚠️  Critical columns not found: {missing_critical}")
        print("   Available columns:", list(df.columns))
        print("   → Update COLUMN_ALIASES at the top of this script to match your column names.")
        return

    print("\n⚙️  Scoring funds...")
    df_scored = score_funds(df, mapping)

    print("\n📝 Building Excel output...")
    build_excel(df_scored, mapping, OUTPUT_FILE)

    print("\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    print(f"  📊 MF Ranked Screener → {OUTPUT_FILE}")
    print(f"  📁 Open this file to see:")
    print(f"     • 🏆 SUMMARY tab — top fund per category at a glance")
    print(f"     • One tab per fund category — all schemes ranked")
    print(f"     • Signals: ⭐ Strong Buy  ✅ Buy  ⚠️ Watch  ❌ Avoid")
    print("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")


if __name__ == "__main__":
    main()
